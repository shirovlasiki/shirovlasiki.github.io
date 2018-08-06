import openpyxl

shironomics = openpyxl.load_workbook('balance.xlsx')
childrenList = shironomics['ChildrenList']
eventList = shironomics['Events']


def find_in_col(ws, col, val):
    for cell in ws[col]:
        if val == cell.value:
            return ws[cell.row]


def text_find_in_col(ws, col, val):
    for cell in ws[col]:
        if cell.value.startswith(val):
            return ws[cell.row]


def set_children_balance(id_num, new_balance):
    find_in_col(childrenList, 'A', id_num)[2].value = new_balance
    shironomics.save('balance.xlsx')


def get_children_balance(id_num):
    return find_in_col(childrenList, 'A', id_num)[2].value


def id_by_name(name):
    return text_find_in_col(childrenList, 'B', name)[0].value


def name_by_id(id_num):
    return find_in_col(childrenList, 'A', id_num)[1].value


def get_children_list():
    cl = []
    keys = []
    for cell in childrenList[1]:
        if len(str(cell.value))>0:
            keys.append(str(cell.value))
        else:
            break
    for cell in childrenList['A']:
        if str(cell.value).isnumeric():
            child = {}
            for i in range(0,len(keys)):
                child[keys[i]] = childrenList[cell.row][i].value
            cl.append(child)

    return cl


def get_event_list():
    el=[]
    for cell in eventList['B']:
        if str(cell.value).isnumeric():
            el.append({'description':eventList[cell.row][0].value, 'price':eventList[cell.row][1].value})
    return el